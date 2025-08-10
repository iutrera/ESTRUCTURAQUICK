"""Genera los nodos de los cuatro cuadrantes aplicando simetrías y desplazamientos."""

import pandas as pd
from sympy import symbols, sympify, Symbol
from openpyxl import load_workbook

# ======================
# VARIABLES SIMBÓLICAS
# ======================

# Definimos variables simbólicas que representan las dimensiones físicas del modelo
WIDTH_CELL = Symbol("WIDTH_CELL")
LENGTH_CELL = Symbol("LENGTH_CELL")
NBR_CELL_WIDTH = Symbol("NBR_CELL_WIDTH")
NBR_CELL_LENGTH = Symbol("NBR_CELL_LENGTH")

# Desplazamientos reales entre cuadrantes
# ⚠️ IMPORTANTE: el desplazamiento es (N - 1) * tamaño_celda
# porque solo hay N-1 saltos entre centros de N celdas
X_OFFSET = (NBR_CELL_WIDTH - 1) * WIDTH_CELL
Y_OFFSET = (NBR_CELL_LENGTH - 1) * LENGTH_CELL

# ======================
# FUNCIÓN SEGURA DE PARSE
# ======================

def parse_expr_safe(expr_str):
    """Convierte texto en expresión simbólica sin evaluarlo."""
    return sympify(expr_str, evaluate=False)

# ======================
# CARGA DE DATOS
# ======================

# Leemos coin1.csv que contiene los nodos base con sus fórmulas simbólicas
coin1_df = pd.read_csv("coin1.csv", dtype={"N": int})
coin1_dict = {
    int(row["N"]): {
        "X": row["X"],
        "Y": row["Y"],
        "Z": row["Z"]
    }
    for _, row in coin1_df.iterrows()
}

# Leemos la hoja 'Correspondance' que contiene el mapeo entre nodos de los 4 cuadrantes
wb = load_workbook("Estructura2.xlsm", data_only=False)
correspondance = wb["Correspondance"]
rows = list(correspondance.iter_rows(min_row=2, max_row=1000, values_only=True))

# Creamos los mapeos: CoinX_ID → Coin1_ID
coin2_map = {int(r[0]): int(r[1]) for r in rows if isinstance(r[0], int) and isinstance(r[1], int)}
coin3_map = {int(r[3]): int(r[4]) for r in rows if isinstance(r[3], int) and isinstance(r[4], int)}
coin4_map = {int(r[6]): int(r[7]) for r in rows if isinstance(r[6], int) and isinstance(r[7], int)}

# ======================
# TRANSFORMACIONES SIMBÓLICAS
# ======================

def apply_coin1(nodo_id):
    """Coin1 = esquina inferior izquierda (sin simetrías ni desplazamientos)"""
    base = coin1_dict.get(nodo_id)
    return {
        "N": nodo_id,
        "X": base["X"],
        "Y": base["Y"],
        "Z": base["Z"]
    }

def apply_coin2(new_id, base_id):
    """Coin2 = esquina superior izquierda (simetría Y + desplazamiento vertical)"""
    base = coin1_dict.get(base_id)
    if base is None:
        return {"N": new_id, "X": "UNDEFINED", "Y": "UNDEFINED", "Z": "UNDEFINED"}
    try:
        x_expr = parse_expr_safe(base["X"])
        y_expr = sympify(f"LENGTH_CELL - ({base['Y']})", evaluate=False) + Y_OFFSET
        z_expr = parse_expr_safe(base["Z"])
        return {"N": new_id, "X": str(x_expr), "Y": str(y_expr), "Z": str(z_expr)}
    except:
        return {"N": new_id, "X": "UNDEFINED", "Y": "UNDEFINED", "Z": "UNDEFINED"}

def apply_coin3(new_id, base_id):
    """Coin3 = esquina inferior derecha (simetría X + desplazamiento horizontal)"""
    base = coin1_dict.get(base_id)
    if base is None:
        return {"N": new_id, "X": "UNDEFINED", "Y": "UNDEFINED", "Z": "UNDEFINED"}
    try:
        x_expr = sympify(f"WIDTH_CELL - ({base['X']})", evaluate=False) + X_OFFSET
        y_expr = parse_expr_safe(base["Y"])
        z_expr = parse_expr_safe(base["Z"])
        return {"N": new_id, "X": str(x_expr), "Y": str(y_expr), "Z": str(z_expr)}
    except:
        return {"N": new_id, "X": "UNDEFINED", "Y": "UNDEFINED", "Z": "UNDEFINED"}

def apply_coin4(new_id, coin2_id):
    """Coin4 = esquina superior derecha (simetría X + Y + ambos desplazamientos)"""
    coin1_id = coin2_map.get(coin2_id)
    base = coin1_dict.get(coin1_id)
    if base is None:
        return {"N": new_id, "X": "UNDEFINED", "Y": "UNDEFINED", "Z": "UNDEFINED"}
    try:
        x_expr = sympify(f"WIDTH_CELL - ({base['X']})", evaluate=False) + X_OFFSET
        y_expr = sympify(f"LENGTH_CELL - ({base['Y']})", evaluate=False) + Y_OFFSET
        z_expr = parse_expr_safe(base["Z"])
        return {"N": new_id, "X": str(x_expr), "Y": str(y_expr), "Z": str(z_expr)}
    except:
        return {"N": new_id, "X": "UNDEFINED", "Y": "UNDEFINED", "Z": "UNDEFINED"}

# ======================
# GENERAR Y GUARDAR CSV
# ======================

# Generamos la lista completa de nodos con sus fórmulas simbólicas ajustadas
coin1_data = [apply_coin1(n) for n in coin1_dict.keys()]
coin2_data = [apply_coin2(n, base) for n, base in coin2_map.items()]
coin3_data = [apply_coin3(n, base) for n, base in coin3_map.items()]
coin4_data = [apply_coin4(n, coin2_id) for n, coin2_id in coin4_map.items()]

# Concatenamos y guardamos
final_df = pd.DataFrame(coin1_data + coin2_data + coin3_data + coin4_data)
final_df.to_csv("nodos_esquinas_simbolico.csv", index=False)
print("✅ Archivo generado: nodos_esquinas_simbolico.csv")
