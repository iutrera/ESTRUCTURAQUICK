"""Convierte fórmulas de Excel a expresiones evaluables y genera `coin1.csv`."""

import re
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

# ==========================
# Entradas por defecto
# ==========================
default_inputs = {
    'LENGTH_CELL': 16.16,
    'WIDTH_CELL': 13.605,
    'LEVEL_FAN_DECK': 26.56,
    'FIRST_LEVEL': 8.85,
    'FAN_DIAMETER': 11.7,
    'NBR_CELL_LENGTH': 6,
    'NBR_CELL_WIDTH': 5,
    'ALFA': 60
}

# ==========================
# Fórmulas derivadas
# ==========================
var_formulas = {
    'Z0': '0.0',
    'Z5': '1.19',
    'Z6': '1.58',
    'Z11': '0.15',
    'X1': '-1.4',
    'X2': '0',
    'X5': '1.23',
    'Y1': '-1.6',
    'Y2': '-0.35',
    'Y3': '0',
    'Y5': '64.1',
    'M1': '1.7',
    'M2': '0.9',
    'M3': '0.4',
    'M4': '6.075',
    'M5': '0.9',
    'M6': '0.3',
    'F11': '0.315',
    'Z1': 'FIRST_LEVEL',
    'Z2': 'FIRST_LEVEL * 2',
    'Z4': 'LEVEL_FAN_DECK - 0.2',
    'Z3': 'Z4 + X1',
    'Z10': '(WIDTH_CELL / 2) / TAN(RADIANS(ALFA) / 2.0)',
    'Z8': 'Z5 + 2.2',
    'Z9': '(Z10 - Z11) / 2',
    'Z7': 'Z9 / 2.0',
    'X3': 'Z9 * TAN(RADIANS(ALFA / 2.0))',
    'X7': 'X3',
    'X4': '(WIDTH_CELL - X3 - X7 - X5) / 2.0',
    'X6': 'X4',
    'X9': 'X5',
    'X8': 'X7 - X9 / 2.0',
    'Y4': 'LENGTH_CELL',
    'R1': 'FAN_DIAMETER / 2.0',
    'F1': '(((LENGTH_CELL - WIDTH_CELL) / 2 + 2 * ((WIDTH_CELL / 2) - ((R1 + F11) / 1.414214))) / 3.0)',
    'F2': 'F1',
    'F3': 'F2',
    'F5': 'F1',
    'F6': 'F2',
    'F4': 'WIDTH_CELL / 2 - F1 - F2 - F3',
    'F7': 'F3',
    'F8': 'LENGTH_CELL / 2 - F5 - F6 - F7',
    'F9': '(LENGTH_CELL - FAN_DIAMETER) / 2 - F11',
    'F10': 'F1 + F2 + F3 - F12',
    'F12': '(WIDTH_CELL - FAN_DIAMETER) / 2 - F11',
    'F13': 'F1 + F2 + F3 - F9',
    'R2': 'R1 + F11',
    'R3': 'R2 + F3 / 1.41421'
}


# ==========================
# Utilidades
# ==========================
def radians(x):
    """Conversión de grados a radianes para usar en `eval`."""
    return math.radians(x)

def eval_expr(expr, context):
    """Evalúa una expresión con las funciones matemáticas permitidas."""
    return eval(expr, {
        "COS": math.cos,
        "SIN": math.sin,
        "TAN": math.tan,
        "PI": math.pi,
        "RADIANS": radians
    }, context)

def build_mapping(ws):
    """Construye un dict con los nombres de variables definidos en la hoja."""
    mapping = {}
    for row in range(1, ws.max_row + 1):
        for name_col, val_col in [('H', 'I'), ('K', 'L'), ('N', 'O'), ('Q', 'R'), ('T', 'U')]:
            name = ws[f"{name_col}{row}"].value
            val = ws[f"{val_col}{row}"].value
            if name and val is not None:
                mapping[f"{val_col}{row}"] = str(name).strip().upper()
        if row == 1 and ws["W1"].value == "α" and ws["X1"].value is not None:
            mapping["X1"] = "ALFA"
        if row >= 2:
            name = ws[f"W{row}"].value
            val = ws[f"X{row}"].value
            if name and val is not None:
                mapping[f"X{row}"] = str(name).strip().upper()
    return mapping

def expand_sum_ranges(expr, ws):
    """Reemplaza rangos de SUMA(A1:A3) por A1+A2+A3."""
    expr_upper = expr.upper()
    keywords = ["SUMA(", "SUM("]
    i = 0
    while i < len(expr_upper):
        for kw in keywords:
            if expr_upper.startswith(kw, i):
                start = i + len(kw)
                level = 1
                end = start
                while end < len(expr_upper) and level > 0:
                    if expr_upper[end] == '(':
                        level += 1
                    elif expr_upper[end] == ')':
                        level -= 1
                    end += 1
                if level != 0:
                    break
                inner = expr[start:end-1]
                args = []
                buf = []
                depth = 0
                for ch in inner:
                    if ch == ',' and depth == 0:
                        args.append(''.join(buf))
                        buf = []
                    else:
                        if ch == '(':
                            depth += 1
                        elif ch == ')':
                            depth -= 1
                        buf.append(ch)
                args.append(''.join(buf))

                expanded_parts = []
                for arg in args:
                    arg_clean = arg.replace('$', '')
                    if ':' in arg_clean:
                        left, right = arg_clean.split(':')
                        try:
                            min_col, min_row, max_col, max_row = range_boundaries(f"{left}:{right}")
                            terms = [f"{get_column_letter(c)}{r}"
                                     for r in range(min_row, max_row+1)
                                     for c in range(min_col, max_col+1)]
                            expanded_parts.append("+".join(terms))
                        except:
                            expanded_parts.append(arg_clean)
                    else:
                        expanded_parts.append(arg_clean)
                expanded = "(" + "+".join(expanded_parts) + ")"
                expr = expr[:i] + expanded + expr[end:]
                expr_upper = expr.upper()
                i += len(expanded) - 1
                break
        i += 1
    return expr

def remove_round(expr):
    """Elimina funciones REDONDEAR/ROUND manteniendo sólo el primer argumento."""
    return re.sub(r"(ROUND|REDONDEAR)\(([^,]+?),\s*\d+\)", r"\2", expr, flags=re.IGNORECASE)

def replace_functions(expr):
    """Normaliza nombres de funciones y separadores al estilo de Python."""
    expr = expr.replace(";", ",")
    expr = re.sub(r"\bPI\(\)", "PI", expr, flags=re.IGNORECASE)
    expr = re.sub(r"\bPI\b", "PI", expr, flags=re.IGNORECASE)
    expr = re.sub(r"\bCOS\(([^()]+)\)", r"COS(\1)", expr, flags=re.IGNORECASE)
    expr = re.sub(r"\bSEN\(([^()]+)\)", r"SIN(\1)", expr, flags=re.IGNORECASE)
    expr = re.sub(r"\bSIN\(([^()]+)\)", r"SIN(\1)", expr, flags=re.IGNORECASE)
    expr = re.sub(r"\bTAN\(([^()]+)\)", r"TAN(\1)", expr, flags=re.IGNORECASE)
    return expr

def replace_cell_refs(expr, mapping):
    """Reemplaza referencias de celda por los nombres definidos en `mapping`."""
    refs = set(re.findall(r"\$?[A-Z]{1,3}\$?\d+", expr))
    for ref in refs:
        clean = ref.replace('$', '')
        if clean in mapping:
            expr = expr.replace(ref, mapping[clean])
    return expr

def clean_formula(expr, mapping, ws):
    """Convierte una fórmula de Excel en una expresión compatible con Python."""
    expr = expr.lstrip('=')
    expr = remove_round(expr)
    expr = expand_sum_ranges(expr, ws)
    expr = replace_functions(expr)
    expr = replace_cell_refs(expr, mapping)
    return expr.upper()

# ==========================
# PROCESO PRINCIPAL
# ==========================
def generar_coin1_csv():
    """Parsea la hoja 'Coin1' y genera el archivo `coin1.csv`."""
    wb = load_workbook("Estructura2.xlsm", data_only=False)
    ws = wb["Coin1"]
    mapping = build_mapping(ws)

    context = default_inputs.copy()

    # Primero evaluamos las variables derivadas
    pendientes = var_formulas.copy()
    while pendientes:
        for var in list(pendientes):
            try:
                context[var] = eval_expr(pendientes[var], context)
                del pendientes[var]
            except:
                continue

    rows = []
    for row in range(2, ws.max_row + 1):
        n = ws[f"A{row}"].value
        if n is None:
            break
        print(f"\n🔹 Nodo {n}")
        record = {"N": n}
        for col, key in [('B', 'X'), ('C', 'Y'), ('D', 'Z')]:
            val = ws[f"{col}{row}"].value
            if isinstance(val, str) and val.startswith('='):
                try:
                    parsed = clean_formula(val, mapping, ws)
                    result = eval_expr(parsed, context)
                    print(f"  {key}:")
                    print(f"    Fórmula original: {val}")
                    print(f"    Fórmula parseada: {parsed}")
                    print(f"    Resultado evaluado: {result}")
                    record[key] = parsed
                except Exception as e:
                    print(f"❌  {key} → Error: {e}")
                    record[key] = f"ERROR({e})"
            else:
                print(f"  {key}: valor directo = {val}")
                record[key] = str(val).upper()
        input("⏩ Pulsa Enter para continuar...")
        rows.append(record)

    df = pd.DataFrame(rows, columns=["N", "X", "Y", "Z"])
    df.to_csv("coin1.csv", index=False, encoding="utf-8")
    print("✅ coin1.csv generado correctamente.")
    return df

if __name__ == "__main__":
    generar_coin1_csv()
